#!/usr/bin/env python3
"""
qa_agent.py — Agente Q&A che risponde citando i paragrafi con anchor [¶N].

Due modalità:
  - full (default):   passa l'intero document.md nel contesto dell'LLM.
                       Il documento è piccolo? L'LLM lo legge tutto e cita.
  - bm25:              indicizza i paragrafi con BM25, recupera i top-K,
                       li passa all'LLM. Per documenti che non entrano nel contesto.

Utilizzo:
  python qa_agent.py --doc-dir processed_documents/subset40
  python qa_agent.py --doc-dir processed_documents/subset40 --mode bm25
  python qa_agent.py --doc-dir processed_documents/subset40 --model llama3
"""
import argparse
import json
import math
import re
import sys
import textwrap
from datetime import datetime
from pathlib import Path
from typing import Any

from datapizza.clients.openai_like import OpenAILikeClient

# ---------------------------------------------------------------------------
# BM25 (implementazione minimale, nessuna dipendenza esterna)
# ---------------------------------------------------------------------------

class BM25:
    """BM25 retriever con parametri standard (k1=1.5, b=0.75)."""

    def __init__(self, k1: float = 1.5, b: float = 0.75):
        self.k1 = k1
        self.b = b
        self._docs: list[list[str]] = []
        self._doc_len: list[int] = []
        self._avgdl: float = 0
        self._df: dict[str, int] = {}   # document frequency
        self._idf: dict[str, float] = {}
        self._N: int = 0

    def tokenize(self, text: str) -> list[str]:
        return re.findall(r"\w+", text.lower())

    def index(self, documents: list[dict[str, Any]], content_key: str = "content"):
        self._docs = [self.tokenize(doc[content_key]) for doc in documents]
        self._N = len(self._docs)
        self._doc_len = [len(d) for d in self._docs]
        self._avgdl = sum(self._doc_len) / max(1, self._N)

        self._df.clear()
        for doc_tokens in self._docs:
            seen = set(doc_tokens)
            for token in seen:
                self._df[token] = self._df.get(token, 0) + 1

        self._idf.clear()
        for token, freq in self._df.items():
            self._idf[token] = math.log((self._N - freq + 0.5) / (freq + 0.5) + 1.0)

    def search(self, query: str, top_k: int = 5) -> list[tuple[int, float]]:
        query_tokens = self.tokenize(query)
        scores: list[float] = [0.0] * self._N

        for token in query_tokens:
            idf = self._idf.get(token, 0)
            if idf == 0:
                continue
            for i, doc_tokens in enumerate(self._docs):
                tf = doc_tokens.count(token)
                if tf == 0:
                    continue
                doc_len_norm = 1 - self.b + self.b * (self._doc_len[i] / self._avgdl)
                score = idf * (tf * (self.k1 + 1)) / (tf + self.k1 * doc_len_norm)
                scores[i] += score

        ranked = sorted(enumerate(scores), key=lambda x: x[1], reverse=True)
        return [(idx, s) for idx, s in ranked[:top_k] if s > 0]


# ---------------------------------------------------------------------------
# Agente Q&A
# ---------------------------------------------------------------------------

SYSTEM_PROMPT_FULL = textwrap.dedent("""\
    Sei un assistente specializzato nell'analisi di documenti tecnici.
    Rispondi alla domanda dell'utente basandoti ESCLUSIVAMENTE sul documento fornito.

    FORMATO CITAZIONI:
    - Ogni paragrafo del documento è seguito da un anchor nel formato [¶N] (es. [¶42]).
    - DEVI citare l'anchor di ogni paragrafo da cui prendi un'informazione.
    - Le citazioni vanno messe DOPO la frase che le usa. Esempio:
      "Il treno deve fermarsi entro 50 metri [¶42]."
    - Se usi più paragrafi, cita ciascuno separatamente.
    - Se il documento non contiene la risposta, dillo esplicitamente.

    FORMATO RISPOSTA:
    - Risposta chiara e concisa.
    - Ogni affermazione deve avere il suo [¶N].
    - Alla fine, elenca le fonti in formato compatto:
      "Fonti: ¶42, ¶58, ¶103"
    """)

SYSTEM_PROMPT_BM25 = textwrap.dedent("""\
    Sei un assistente specializzato nell'analisi di documenti tecnici.
    Rispondi alla domanda dell'utente basandoti ESCLUSIVAMENTE sui paragrafi forniti.

    FORMATO CITAZIONI:
    - Ogni paragrafo ha un anchor [¶N] e metadati (pagina, tipo).
    - DEVI citare l'anchor di ogni paragrafo da cui prendi un'informazione.
    - Esempio: "Il treno deve fermarsi entro 50 metri [¶42]."
    - Se i paragrafi forniti non contengono la risposta, dillo esplicitamente.

    FORMATO RISPOSTA:
    - Risposta chiara e concisa.
    - Ogni affermazione deve avere il suo [¶N].
    - Alla fine, elenca le fonti: "Fonti: ¶42, ¶58"
    """)


class QAAgent:
    """Agente Q&A con citazione dei paragrafi tramite anchor Docling [¶N].

    Supporta uno o più subset di documenti. Con più subset, gli anchor
    vengono prefissati con il numero del subset (es. ¶35-1, ¶40-1, ¶58-1).
    """

    def __init__(
        self,
        doc_dirs: list[str],
        model: str = "deepseek-v4-flash:cloud",
        api_base: str = "http://localhost:11434/v1",
        api_key: str = "ollama",
    ):
        self.doc_dirs = [Path(d) for d in doc_dirs]
        self.model = model
        self.api_base = api_base
        self.api_key = api_key

        self._citations: dict[str, dict] = {}
        self._citable_chunks: list[dict] = []
        self._markdown: str = ""
        self._client: OpenAILikeClient | None = None
        self._anchor_to_doc_dir: dict[str, Path] = {}
        # Primo doc_dir per retrocompatibilità
        self.doc_dir = self.doc_dirs[0]

    # ── Inizializzazione ─────────────────────────────────────────────────

    def load(self) -> "QAAgent":
        """Carica document.md e citations.json da tutti i subset."""
        md_parts: list[str] = []
        multi = len(self.doc_dirs) > 1

        for doc_dir in self.doc_dirs:
            md_path = doc_dir / "document.md"
            citations_path = doc_dir / "citations.json"

            if not md_path.exists():
                raise FileNotFoundError(f"Markdown non trovato: {md_path}")
            if not citations_path.exists():
                raise FileNotFoundError(f"Citations non trovato: {citations_path}")

            md_text = md_path.read_text(encoding="utf-8")
            raw = json.loads(citations_path.read_text(encoding="utf-8"))
            subset_citations = raw.get("citations", {})

            if multi:
                prefix = re.sub(r"(?i)^subset[_-]?", "", doc_dir.name).replace(" v360", "")
                # Rinomina anchor nel markdown: [¶N] → [¶{prefix}-N]
                md_text = re.sub(r"\[¶(\d+)\]", rf"[¶{prefix}-\1]", md_text)
                # Rinomina chiavi nelle citations
                subset_citations = {
                    f"¶{prefix}-{k.lstrip('¶')}": v
                    for k, v in subset_citations.items()
                }

            md_parts.append(md_text)
            self._citations.update(subset_citations)

            # Mappa anchor → doc_dir per i link al viewer
            for anchor in subset_citations:
                self._anchor_to_doc_dir[anchor] = doc_dir

        self._markdown = "\n\n".join(md_parts)

        # Lista ordinata di chunk citabili (per BM25)
        self._citable_chunks = [
            {"anchor": anchor, **meta}
            for anchor, meta in self._citations.items()
        ]
        self._citable_chunks.sort(key=lambda c: c["anchor"])

        self._init_client()
        return self

    def _init_client(self):
        self._client = OpenAILikeClient(
            api_key=self.api_key,
            model=self.model,
            base_url=self.api_base,
        )

    # ── Chiamata LLM ────────────────────────────────────────────────────

    def _call_llm(self, system_prompt: str, user_prompt: str) -> tuple[str, dict[str, int]]:
        """Restituisce (testo_risposta, {prompt_tokens, completion_tokens, total_tokens})."""
        response = self._client.invoke(
            input=user_prompt,
            system_prompt=system_prompt,
            temperature=0.1,
        )
        text = response.text or ""
        usage = {
            "prompt_tokens": response.usage.prompt_tokens,
            "completion_tokens": response.usage.completion_tokens,
            "total_tokens": response.usage.prompt_tokens + response.usage.completion_tokens,
        }
        return text, usage

    # ── Lookup fonti ────────────────────────────────────────────────────

    def _extract_anchors(self, text: str) -> list[str]:
        """Estrae tutti gli anchor ¶N o ¶{prefix}-N unici dal testo, ordinati."""
        found = re.findall(r"¶(?:(\d+)-)?(\d+)", text)
        seen: set[str] = set()
        unique: list[str] = []
        for prefix, num in found:
            anchor = f"¶{prefix}-{num}" if prefix else f"¶{num}"
            if anchor not in seen and anchor in self._citations:
                seen.add(anchor)
                unique.append(anchor)
        return unique

    def _lookup_sources(self, anchors: list[str]) -> list[dict]:
        result = []
        for a in anchors:
            c = self._citations[a]
            doc_dir = self._anchor_to_doc_dir.get(a, self.doc_dir)
            source = {
                "anchor": a,
                "page": c["page_id"],
                "type": c["type"],
                "content": c["content"],
                "doc_dir": str(doc_dir),
            }
            if c.get("bbox"):
                source["bbox"] = c["bbox"]
            result.append(source)
        return result

    # ── Modalità Full-Context ───────────────────────────────────────────

    def _ask_full(self, question: str) -> dict:
        user_prompt = (
            f"DOCUMENTO:\n\n{self._markdown}\n\n"
            f"DOMANDA: {question}"
        )

        answer, tokens = self._call_llm(SYSTEM_PROMPT_FULL, user_prompt)
        anchors = self._extract_anchors(answer)
        sources = self._lookup_sources(anchors)
        return {"answer": answer, "sources": sources, "mode": "full", "tokens": tokens}

    # ── Modalità BM25 ───────────────────────────────────────────────────

    def _ask_bm25(self, question: str, top_k: int = 10) -> dict:
        bm25 = BM25()
        bm25.index(self._citable_chunks, content_key="content")
        results = bm25.search(question, top_k=top_k)

        if not results:
            return {
                "answer": "Nessun paragrafo rilevante trovato nel documento.",
                "sources": [],
                "mode": "bm25",
                "tokens": {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0},
            }

        # Costruisce il prompt con i paragrafi recuperati
        paragraphs: list[str] = []
        retrieved_anchors: list[str] = []
        for idx, score in results:
            chunk = self._citable_chunks[idx]
            anchor = chunk["anchor"]
            retrieved_anchors.append(anchor)
            meta = self._citations[anchor]
            paragraphs.append(
                f"[{anchor}] (pagina {meta['page_id']}, {meta['type']}):\n"
                f"{meta['content']}\n"
            )

        context_block = "\n".join(paragraphs)
        user_prompt = (
            f"PARAGRAFI RILEVANTI:\n\n{context_block}\n\n"
            f"DOMANDA: {question}"
        )

        answer, tokens = self._call_llm(SYSTEM_PROMPT_BM25, user_prompt)
        anchors = self._extract_anchors(answer)
        sources = self._lookup_sources(anchors)
        return {"answer": answer, "sources": sources, "mode": f"bm25 (top {len(results)})", "tokens": tokens}

    # ── API pubblica ────────────────────────────────────────────────────

    def ask(self, question: str, mode: str = "full", top_k: int = 10) -> dict:
        if mode == "bm25":
            return self._ask_bm25(question, top_k=top_k)
        return self._ask_full(question)


# ---------------------------------------------------------------------------
# Formattazione output
# ---------------------------------------------------------------------------

def format_answer(result: dict) -> str:
    lines = [
        f"{'─' * 60}",
        f"Modalità: {result['mode']}",
        f"{'─' * 60}",
        "",
        result["answer"].rstrip(),
    ]

    if result["sources"]:
        lines.append("")
        lines.append("─" * 60)
        lines.append("FONTI VERIFICATE:")
        lines.append("")
        for i, src in enumerate(result["sources"], 1):
            content_preview = src["content"][:200].replace("\n", " ")
            if len(src["content"]) > 200:
                content_preview += "…"
            lines.append(f"  [{src['anchor']}] pagina {src['page']}, {src['type']}")
            lines.append(f"      {content_preview}")
            lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Elaborazione batch da Excel → report HTML
# ---------------------------------------------------------------------------

def _read_questions(excel_path: str) -> list[dict]:
    """Legge le domande da un foglio Excel.

    Colonne attese:
      A: Domanda
      B: Risposta (attesa, opzionale)
      C: Pagina  (attesa, opzionale)

    Restituisce lista di {idx, question, expected_answer, expected_page}.
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    questions: list[dict] = []
    for row in range(2, ws.max_row + 1):
        question = ws.cell(row, 1).value
        if not question or not str(question).strip():
            continue
        expected_answer = ws.cell(row, 2).value
        expected_page = ws.cell(row, 3).value
        questions.append({
            "idx": row - 2,
            "question": str(question).strip(),
            "expected_answer": str(expected_answer).strip() if expected_answer else None,
            "expected_page": expected_page,
        })

    wb.close()
    return questions


def process_excel_to_html(
    agent: "QAAgent",
    excel_path: str,
    output_path: str | None = None,
    mode: str = "full",
    top_k: int = 10,
) -> Path:
    """Elabora le domande di un Excel e produce un report HTML con fonti cliccabili.

    Ogni citazione [¶N] nella risposta diventa un link che apre
    annotated_bboxes.pdf alla pagina esatta con le bounding box.
    """
    from html_reporter import generate_html_report

    questions = _read_questions(excel_path)
    if not questions:
        print("❌ Nessuna domanda trovata nel file Excel.")
        sys.exit(1)

    total = len(questions)
    print(f"\n📋 Trovate {total} domande in: {excel_path}")
    print(f"   Modalità: {mode}  |  Top-K: {top_k}  |  Modello: {agent.model}")
    print()

    # Genera output path se non specificato
    if output_path is None:
        src = Path(excel_path)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        model_safe = agent.model.replace(":", "_").replace("/", "_")
        output_path = str(src.parent / f"report_{src.stem}_{model_safe}_{ts}.html")

    results: list[dict] = []

    for i, q in enumerate(questions):
        print(f"[{i + 1}/{total}] {q['question'][:100]}{'…' if len(q['question']) > 100 else ''}")

        result = agent.ask(q["question"], mode=mode, top_k=top_k)
        results.append({
            "question": q["question"],
            "expected_answer": q["expected_answer"],
            "expected_page": q["expected_page"],
            "answer": result["answer"].rstrip(),
            "sources": result["sources"],
            "tokens": result.get("tokens", {}),
        })

        preview = result["answer"][:150].replace("\n", " ")
        if len(result["answer"]) > 150:
            preview += "…"
        print(f"   → {preview}")
        if result["sources"]:
            fonti = ", ".join(f"{s['anchor']} (p.{s['page']})" for s in result["sources"])
            print(f"   → Fonti: {fonti}")
        print()

    doc_dirs_list = [str(d) for d in agent.doc_dirs]
    title = f"Q&A: {', '.join(d.name for d in agent.doc_dirs)}"
    html_path = generate_html_report(results, doc_dirs_list, output_path, model=agent.model, title=title)
    print(f"✅ Report HTML salvato in: {html_path}")
    return html_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Agente Q&A con citazione dei paragrafi via anchor Docling [¶N].",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            esempi:
              python qa_agent.py --doc-dir processed_documents/subset40
              python qa_agent.py -d processed_documents/subset35 processed_documents/subset40 processed_documents/subset58
              python qa_agent.py -d processed_documents/subset40 -m bm25
              python qa_agent.py -d processed_documents/subset40 -q "quanto deve fermarsi il treno?"
              python qa_agent.py -d processed_documents/subset35 processed_documents/subset40 processed_documents/subset58 --excel "domande_excel/Domande Subset-35-40-58.xlsx"
        """),
    )
    p.add_argument("-d", "--doc-dir", required=True, nargs="+",
                   help="Cartella/e con document.md e citations.json (una o più)")
    p.add_argument("-m", "--mode", choices=["full", "bm25"], default="full",
                   help="Modalità: full (default) o bm25 per documenti grandi")
    p.add_argument("-q", "--question", default=None,
                   help="Domanda singola (senza --excel)")
    p.add_argument("--excel", default=None,
                   help="File Excel con domande da elaborare in batch")
    p.add_argument("--output", default=None,
                   help="Percorso file HTML di output (default: generato automaticamente)")
    p.add_argument("--model", default="deepseek-v4-flash:cloud",
                   help="Modello LLM (default: deepseek-v4-flash:cloud)")
    p.add_argument("--api-base", default="http://localhost:11434/v1",
                   help="URL API OpenAI-compatibile")
    p.add_argument("--api-key", default="ollama",
                   help="API key")
    p.add_argument("--top-k", type=int, default=10,
                   help="Paragrafi da recuperare in modalità bm25 (default: 10)")
    p.add_argument("--no-color", action="store_true",
                   help="Disabilita codici colore ANSI")
    return p.parse_args()


def _expand_doc_dirs(doc_dirs: list[str]) -> list[str]:
    """Se una directory è la root 'processed_documents', la espande in tutte le sottocartelle."""
    expanded: list[str] = []
    for d in doc_dirs:
        p = Path(d).resolve()
        # Se la cartella contiene già document.md, la usiamo così com'è
        if (p / "document.md").exists():
            expanded.append(d)
        else:
            # Altrimenti cerchiamo le sottocartelle che contengono document.md
            subs = sorted(
                str(sub) for sub in p.iterdir()
                if sub.is_dir() and (sub / "document.md").exists()
            )
            if not subs:
                raise FileNotFoundError(
                    f"Né {d} né le sue sottocartelle contengono document.md"
                )
            expanded.extend(subs)
    return expanded


def main() -> None:
    args = parse_args()
    args.doc_dir = _expand_doc_dirs(args.doc_dir)

    try:
        agent = QAAgent(
            doc_dirs=args.doc_dir,
            model=args.model,
            api_base=args.api_base,
            api_key=args.api_key,
        ).load()
    except FileNotFoundError as e:
        print(f"❌ {e}")
        sys.exit(1)

    if args.excel:
        process_excel_to_html(
            agent=agent,
            excel_path=args.excel,
            output_path=args.output,
            mode=args.mode,
            top_k=args.top_k,
        )
    elif args.question:
        result = agent.ask(args.question, mode=args.mode, top_k=args.top_k)
        print(format_answer(result))
    else:
        # REPL interattivo
        print(f"\n🤖 Q&A Agent — subset: {', '.join(d.name for d in agent.doc_dirs)}")
        print(f"   Modello: {agent.model}  |  Modalità: {args.mode}")
        print(f"   Documento combinato: {len(agent._markdown):,} char")
        print(f"   Paragrafi citabili: {len(agent._citations)}")
        print("\n   Scrivi 'fine' o 'quit' per uscire, 'mode full|bm25' per cambiare.\n")

        mode = args.mode
        while True:
            try:
                question = input("❯ ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\n")
                break

            if not question:
                continue

            if question.lower() in ("fine", "quit", "exit", "q"):
                break

            if question.lower().startswith("mode "):
                new_mode = question.split()[1]
                if new_mode in ("full", "bm25"):
                    mode = new_mode
                    print(f"   → modalità: {mode}")
                else:
                    print(f"   → modalità sconosciuta: {new_mode}")
                continue

            result = agent.ask(question, mode=mode, top_k=args.top_k)
            print(format_answer(result))
            print()


if __name__ == "__main__":
    main()
